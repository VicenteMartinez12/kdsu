import os
import magic  # python-magic
from openpyxl import load_workbook

def read_xlsx(file_path: str):
    """
    Lee un archivo XLSX (solo la primera hoja) y lo convierte en una lista de diccionarios.
    :param file_path: Ruta del archivo XLSX
    :return: Lista de diccionarios con los datos de la hoja
    :raises FileNotFoundError: Si el archivo no existe
    :raises ValueError: Si el archivo no es un XLSX válido
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"El archivo '{file_path}' no existe.")

    mime = magic.from_file(file_path, mime=True)
    if mime != 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        raise ValueError(f"El archivo '{file_path}' no es un archivo XLSX válido. MIME: {mime}")

    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active  # Toma la primera hoja activa

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = rows[0]  # Primera fila como encabezados
    data = []
    for row in rows[1:]:
        row_dict = dict(zip(headers, row))
        data.append(row_dict)

    return data
